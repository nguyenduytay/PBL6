"""
Export endpoints - API xuất dữ liệu phân tích ra CSV/JSON/Excel
"""
import os
import sys
import csv
import json
from io import BytesIO, StringIO
from typing import List
from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import Response

# Thêm project root vào path
project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..', '..'))
if project_root not in sys.path:
    sys.path.insert(0, project_root)

from app.services.analysis_service import AnalysisService

router = APIRouter()
analysis_service = AnalysisService()  # Service lấy dữ liệu analyses


@router.get("/analyses/csv")
async def export_analyses_csv(
    limit: int = Query(1000, ge=1, le=10000),
    offset: int = Query(0, ge=0)
):
    """
    Export analyses ra CSV
    
    - limit: Số lượng records (tối đa 10000)
    - offset: Vị trí bắt đầu
    """
    try:
        analyses = await analysis_service.get_all(limit=limit, offset=offset)
        
        if not analyses:
            raise HTTPException(status_code=404, detail="No analyses found")
        
        # Tạo file CSV với các cột cần thiết
        output = StringIO()
        writer = csv.DictWriter(output, fieldnames=[
            'id', 'filename', 'sha256', 'md5', 'file_size', 
            'malware_detected', 'analysis_time', 'created_at'
        ])
        writer.writeheader()  # Ghi header
        
        # Ghi từng dòng dữ liệu
        for analysis in analyses:
            writer.writerow({
                'id': analysis.get('id'),
                'filename': analysis.get('filename'),
                'sha256': analysis.get('sha256') or '',
                'md5': analysis.get('md5') or '',
                'file_size': analysis.get('file_size') or 0,
                'malware_detected': 'Yes' if analysis.get('malware_detected') else 'No',
                'analysis_time': analysis.get('analysis_time', 0.0),
                'created_at': str(analysis.get('created_at', ''))
            })
        
        csv_content = output.getvalue()
        output.close()
        
        return Response(
            content=csv_content,
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=analyses_export.csv"}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error exporting CSV: {str(e)}")


@router.get("/analyses/json")
async def export_analyses_json(
    limit: int = Query(1000, ge=1, le=10000),
    offset: int = Query(0, ge=0)
):
    """
    Xuất dữ liệu ra JSON
    
    - limit: Số lượng records (tối đa 10000)
    - offset: Vị trí bắt đầu
    """
    try:
        analyses = await analysis_service.get_all(limit=limit, offset=offset)
        
        if not analyses:
            raise HTTPException(status_code=404, detail="No analyses found")
        
        # Chuyển đổi sang JSON với format đẹp
        json_content = json.dumps(analyses, indent=2, default=str)
        
        return Response(
            content=json_content,
            media_type="application/json",
            headers={"Content-Disposition": "attachment; filename=analyses_export.json"}  # Tên file download
        )
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error exporting JSON: {str(e)}")


@router.get("/analyses/excel")
async def export_analyses_excel(
    limit: int = Query(1000, ge=1, le=10000),
    offset: int = Query(0, ge=0)
):
    """
    Xuất dữ liệu ra file Excel (XLSX)
    
    - limit: Số lượng records (tối đa 10000)
    - offset: Vị trí bắt đầu
    """
    try:
        analyses = await analysis_service.get_all(limit=limit, offset=offset)
        
        if not analyses:
            raise HTTPException(status_code=404, detail="No analyses found")
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            # Tạo workbook Excel
            wb = Workbook()
            ws = wb.active
            ws.title = "Các mẫu phân tích"
            
            # Tạo tiêu đề cột
            headers = ['ID', 'Tên file', 'SHA256', 'MD5', 'Kích thước (bytes)', 
                      'Phát hiện Malware', 'Thời gian phân tích (s)', 'Ngày tạo']
            ws.append(headers)
            
            # Định dạng tiêu đề (màu xanh, chữ trắng, in đậm)
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Thêm dữ liệu vào các dòng
            for analysis in analyses:
                ws.append([
                    analysis.get('id'),
                    analysis.get('filename', ''),
                    analysis.get('sha256', ''),
                    analysis.get('md5', ''),
                    analysis.get('file_size', 0),
                    'Yes' if analysis.get('malware_detected') else 'No',
                    analysis.get('analysis_time', 0.0),
                    str(analysis.get('created_at', ''))
                ])
            
            # Tự động điều chỉnh độ rộng cột theo nội dung
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Tối đa 50 ký tự
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Lưu workbook vào bộ nhớ đệm
            output = BytesIO()
            wb.save(output)
            output.seek(0)  # Reset pointer về đầu file
            
            return Response(
                content=output.getvalue(),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": "attachment; filename=analyses_export.xlsx"}
            )
            
        except ImportError:
            raise HTTPException(
                status_code=500,
                detail="openpyxl not installed. Install with: pip install openpyxl"
            )
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error exporting Excel: {str(e)}")



